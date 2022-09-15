import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import xlrd
import hashlib
from pandas import DataFrame
from collections import Counter
from openpyxl import load_workbook
from openpyxl.styles  import Font

path = 'Q.txt'
f = open(path, 'w',encoding='UTF-8')

a=pd.read_excel('C:/Users/建良/Desktop/co/data.xlsx')
df = pd.DataFrame(a)

sp = str(input("請輸入要修改欄位："))
List= df[sp].tolist()

df1=df[sp].apply(str)
u=[]
for i in df1:
    u.append(i)

T="".join(u)

sha = hashlib.md5(T.encode("utf-8")).hexdigest()
print(sha,file=f)  #要留 存檔

bin_str = ""
for n in sha:
    bin_str += bin(int(n,16))[2:].zfill(4)
print(bin_str,file=f)  #要留 存檔

recounted = Counter(List)   #統計資料出現次數
b1=df.groupby(sp).size()
A=max(b1)  #找出最大值的點為多少


a=np.array(List)   #將資料改成陣列  (分數)
a1=np.unique(a)    

b=np.array(b1)     #將資料轉成陣列  (次數)

c=np.vstack((b,a1))   #合併變成二維陣列

MAX=max(List)

plt.hist(List,range(0,MAX+2),align='left', edgecolor='#000000',linewidth=2)
plt.xticks(List)
plt.savefig("hist0.png")
plt.close()

n=-1
p=[]

for row in c:    
    for col in row: 
        n=n+1
        if(col==A):     #判斷次數如果是跟最大一樣
           p.append(c[1][n])

M=p[0]     #找第一個出現的最大值

m=-1
o=[]
for i in a:
    m=m+1  
    if i < M:  
        i=i-1
        df.at[m,sp] = i
        DataFrame(df).to_excel('C:/Users/建良/Desktop/co/hide.xlsx',sheet_name='Sheet1', index=False, header=True)
        if(i<0): 
            o.append(m)
    else:     
        if(i==M):
            df.at[m,sp] = i
            DataFrame(df).to_excel('C:/Users/建良/Desktop/co/hide.xlsx',sheet_name='Sheet1', index=False, header=True)
        else:
            i=i+1
            df.at[m,sp] = i
            DataFrame(df).to_excel('C:/Users/建良/Desktop/co/hide.xlsx',sheet_name='Sheet1', index=False, header=True)  

def union_without_repetition(list1,list2):
   result = list(set(list1) | set(list2))
   return result

r=pd.read_excel('C:/Users/建良/Desktop/co/hide.xlsx',sheet_name='Sheet1')
sr= df[sp].tolist()
ar=np.array(sr)

bins_listx=union_without_repetition(List,sr)
plt.hist(sr,range(-1,MAX+2),align='left', edgecolor='#000000',linewidth=2)
plt.xticks(bins_listx)
plt.savefig("hist1.png")
plt.close()

tc=len(bin_str)

n=-1
t=-1
for i in ar:
    n=n+1
    if(i==M):
        t=t+1    #計算第在第幾個資料
        if(t<tc):
            if(bin_str[t]=='1'):
                i=i-1
                df.at[n,sp] = i
                DataFrame(df).to_excel('C:/Users/建良/Desktop/co/hide.xlsx',sheet_name='Sheet1', index=False, header=True)
            if(bin_str[t]=='0'):
                DataFrame(df).to_excel('C:/Users/建良/Desktop/co/hide.xlsx',sheet_name='Sheet1', index=False, header=True)
            if(bin_str[t]==' '):
                i=i+1
                df.at[n,sp] = i
                DataFrame(df).to_excel('C:/Users/建良/Desktop/co/hide.xlsx',sheet_name='Sheet1', index=False, header=True)
    else:
        df.at[n,sp]=i
        DataFrame(df).to_excel('C:/Users/建良/Desktop/co/hide.xlsx',sheet_name='Sheet1', index=False, header=True)



def union_without_repetition(list1,list2):
   result = list(set(list1) | set(list2))
   return result

wb=load_workbook(filename = 'hide.xlsx')
sheet_ranges = wb['Sheet1']

k = pd.read_excel("hide.xlsx",nrows=0)
L = k.columns.tolist()
r=L.index(sp)+1

font = Font(size=15)

for i in o:
    i=i+2 #PYTHON讀的格子0 在EXCEL是2
    f = sheet_ranges.cell(row=i, column=r)
    f.font = font
    f.value=0
    wb.save('hide.xlsx')

r=pd.read_excel('C:/Users/建良/Desktop/co/hide.xlsx')
sx= df[sp].tolist()
bins_listx2=union_without_repetition(bins_listx,sx)
plt.hist(sx,range(0,MAX+2),align='left', edgecolor='#000000',linewidth=2)
plt.xticks(bins_listx2)
plt.savefig("hist2.png")

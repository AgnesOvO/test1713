from app import app
from flask import Flask, render_template, request, redirect, send_from_directory, abort, flash, session, Blueprint
import os
from os import listdir
from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook
from openpyxl.styles import Font
from werkzeug.utils import secure_filename
#修改excel
import pandas as pd 
from pandas import DataFrame 
import numpy as np 
import matplotlib.pyplot as plt 
from collections import Counter
import xlrd
import hashlib
#用sqlite上傳檔案
#from . import db
#用MongoDB上傳檔案
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles  import Font
#import pymongo
#from pymongo import MongoClient #使用mongodb
#import certifi #為了解決連線到SSL的問題
import pathlib #分割副檔名
import csv
import shutil #移動檔案 覆蓋原檔案
#import pandas as pd 重複了
import json
#比對雜湊值
import tk
#import tkinter as tk
import filecmp
#set FLASK_ENV=development


@app.route("/") #主頁
def index():
    return render_template("public/index.html")

def allowed_excel(filename):

    #上傳的文件要有副檔名
    if not "." in filename:
        return False

    #將.從副檔名中分割出來
    ext = filename.rsplit(".", 1)[1]

    #確認副檔名和ALLOWED_EXCEL_EXTENSIONS中的一樣
    if ext.upper() in app.config["ALLOWED_EXCEL_EXTENSIONS"]:
        return True
    else:
        return False                

app.config["EXCEL_UPLOADS"] = "/app/app/static/upload/excel_hide/excel_ori" #儲存位置
app.config["ALLOWED_EXCEL_EXTENSIONS"] = ["XLSX"] #允許的副檔名
app.config["SECRET_KEY"] = "OCML3BOswQEUeaxcuKHLpw" #隨機產生的SECRET_KEY，有這個才能跑flash

@app.route("/upload-excel", methods=["GET", "POST"]) #上傳excel檔
def upload_excel():

    if request.method == "POST":

        if request.files:

            excel = request.files["excel"]

            if excel.filename == "":
                flash('未選取檔案', 'warning')
                return redirect(request.url)

            if allowed_excel(excel.filename):
                filename = secure_filename(excel.filename)
                
                #如果檔名已經存在，則刪除舊檔，建立新檔
                if os.path.isfile(app.config["EXCEL_UPLOADS"] + excel.filename):
                    os.remove(app.config["EXCEL_UPLOADS"] + excel.filename)
                    excel.save(os.path.join(app.config["EXCEL_UPLOADS"], excel.filename))
                else:
                    excel.save(os.path.join(app.config["EXCEL_UPLOADS"], excel.filename))

                #如果ori.xlsx存在，則刪除舊檔，建立新檔
                str_upload_path = str(app.config["EXCEL_UPLOADS"])
                if os.path.isfile(app.config["EXCEL_UPLOADS"] + "/ori.xlsx"):
                    os.remove(app.config["EXCEL_UPLOADS"] + "/ori.xlsx")
                    os.rename(str_upload_path + "/" + excel.filename,str_upload_path + "/" + "ori.xlsx")
                else:
                    os.rename(str_upload_path + "/" + excel.filename,str_upload_path + "/" + "ori.xlsx")

                flash('Excel saved', 'success')
                return redirect(request.url)
                #return redirect("/download/"+filename) #會下載剛剛上傳的檔案

            else:
                flash('請上傳附檔名為 .xlsx 的檔案', 'warning')
                return redirect(request.url)

    return render_template("public/data_hiding.html")

#下載檔案，用from flask import send_from_directory, abort
#app.config["CLIENT_EXCELS"] = "/app/app/static/excel" #要從哪裡下載

app.config["EXCEL_SHA"] = "/app/app/static/upload/excel_hide/excel_sha" #雜湊值 存檔路徑 C:\app\app\static\upload\excel_hash
app.config["EXCEL_HASH"] = "/app/app/static/upload/excel_hide/excel_hash" #藏入雜湊值後的檔案 存檔路徑
#app.config["EXCEL_HIST"] = "/app/app/static/excel_hist" #excel產生的直方圖存檔路徑

def popup(MAX):
    flash('請記下來，此檔案的最高值: ' + str(MAX), 'warning')

#輸入欄位名稱_藏入雜湊值
@app.route("/add-hash", methods=["POST"])
def add_hash():
    return render_template("public/addHASH.html")

#藏入雜湊值
@app.route("/HASH", methods=["GET", "POST"])
def hash():

    # 使用openpyxl建立新活頁簿wb_new
    wb_new = Workbook()
    wb_new.save(app.config["EXCEL_HASH"] + '/new_excel_hash.xlsx')

    # 使用openpyxl讀取原始檔案
    wb = load_workbook(app.config["EXCEL_UPLOADS"] + '/ori.xlsx')
    ws = wb.worksheets[0]

    # 使用openpyxl讀取new_excel
    wb_new = load_workbook(app.config["EXCEL_HASH"] + '/new_excel_hash.xlsx')
    ws_new = wb_new.active

    #雜湊值的檔案
    path = app.config["EXCEL_SHA"] + '/sha.txt'
    f = open(path, 'w',encoding='UTF-8')

    #讀舊檔，輸入欄位名稱
    a=pd.read_excel(app.config["EXCEL_UPLOADS"] + '/ori.xlsx')
    df = pd.DataFrame(a)
    sp = request.args.get('HASHcolname') #request.args.get('HASHcolname')是由使用者輸入的欄位名稱
    List= df[sp].tolist()

    df1=df[sp].apply(str)
    u=[]
    for i in df1:
        u.append(i)

    T="".join(u)

    #生成雜湊值
    sha = hashlib.md5(T.encode("utf-8")).hexdigest()
    print(sha,file=f)  #要留 存檔

    bin_str = ""
    for n in sha:
        bin_str += bin(int(n,16))[2:].zfill(4)
    print(bin_str,file=f)  #要留 存檔

    recounted = Counter(List)   #統計資料出現次數
    b1=df.groupby(sp).size()
    A=max(b1)  #找出最大值的點為多少

    a=np.array(List)   #將資料改成陣列  (分數)
    a1=np.unique(a)    

    b=np.array(b1)     #將資料轉成陣列  (次數)

    c=np.vstack((b,a1))   #合併變成二維陣列

    MAX=max(List)

    #plt.hist(List,range(0,MAX+2),align='left', edgecolor='#000000',linewidth=2)
    #plt.xticks(List)
    #plt.savefig(app.config["EXCEL_HIST"] + "/hist_ori.png") #原始資料的直方圖
    #plt.close()

    n=-1
    p=[]

    for row in c:    
        for col in row: 
            n=n+1
            if(col==A):     #判斷次數如果是跟最大一樣
                p.append(c[1][n])

    M=p[0]     #找第一個出現的最大值
    #print('最高點', M)
    m=-1
    o=[]
    for i in a:
        m=m+1  
        if i < M:  
            i=i-1
            df.at[m,sp] = i
            DataFrame(df).to_excel(app.config["EXCEL_HASH"] + '/new_excel_hash.xlsx',sheet_name='Sheet1', index=False, header=True)
            if(i<0): 
                o.append(m)
        else:     
            if(i==M):
                df.at[m,sp] = i
                DataFrame(df).to_excel(app.config["EXCEL_HASH"] + '/new_excel_hash.xlsx',sheet_name='Sheet1', index=False, header=True)
            else:
                i=i+1
                df.at[m,sp] = i
                DataFrame(df).to_excel(app.config["EXCEL_HASH"] + '/new_excel_hash.xlsx',sheet_name='Sheet1', index=False, header=True)  

    def union_without_repetition(list1,list2):
        result = list(set(list1) | set(list2))
        return result

    r=pd.read_excel(app.config["EXCEL_HASH"] + '/new_excel_hash.xlsx',sheet_name='Sheet1')
    sr= df[sp].tolist()
    ar=np.array(sr)

    bins_listx=union_without_repetition(List,sr)
    #plt.hist(sr,range(-1,MAX+2),align='left', edgecolor='#000000',linewidth=2)
    #plt.xticks(bins_listx)
    #plt.savefig(app.config["EXCEL_HIST"] + "/hist_shifting.png")
    #plt.close()

    tc=len(bin_str)

    n=-1
    t=-1
    for i in ar:
        n=n+1
        if(i==M):
            t=t+1    #計算第在第幾個資料
            if(t<tc):
                if(bin_str[t]=='1'):
                    i=i-1
                    df.at[n,sp] = i
                    DataFrame(df).to_excel(app.config["EXCEL_HASH"] + '/new_excel_hash.xlsx',sheet_name='Sheet1', index=False, header=True)
                if(bin_str[t]=='0'):
                    DataFrame(df).to_excel(app.config["EXCEL_HASH"] + '/new_excel_hash.xlsx',sheet_name='Sheet1', index=False, header=True)
                if(bin_str[t]==' '):
                    i=i+1
                    df.at[n,sp] = i
                    DataFrame(df).to_excel(app.config["EXCEL_HASH"] + '/new_excel_hash.xlsx',sheet_name='Sheet1', index=False, header=True)
        else:
            df.at[n,sp]=i
            DataFrame(df).to_excel(app.config["EXCEL_HASH"] + '/new_excel_hash.xlsx',sheet_name='Sheet1', index=False, header=True)



    def union_without_repetition(list1,list2):
        result = list(set(list1) | set(list2))
        return result

    fn = app.config["EXCEL_HASH"] + '/new_excel_hash.xlsx'
    wb=load_workbook(fn)
    sheet_ranges = wb['Sheet1']

    k = pd.read_excel(app.config["EXCEL_HASH"] + '/new_excel_hash.xlsx',nrows=0)
    L = k.columns.tolist()
    r=L.index(sp)+1

    font = Font(size=15)

    for i in o:
        i=i+2 #PYTHON讀的格子0 在EXCEL是2
        f = sheet_ranges.cell(row=i, column=r)
        f.font = font
        f.value=0
        wb.save(app.config["EXCEL_HASH"] + '/new_excel_hash.xlsx')

    r=pd.read_excel(app.config["EXCEL_HASH"] + '/new_excel_hash.xlsx')
    sx= df[sp].tolist()
    bins_listx2=union_without_repetition(bins_listx,sx)
    #plt.hist(sx,range(0,MAX+2),align='left', edgecolor='#000000',linewidth=2)
    #plt.xticks(bins_listx2)
    #plt.savefig(app.config["EXCEL_HIST"] + "/hist_hiding.png")
    if M is not None:
        popup(M) #FLASH最高值
        return render_template("public/addHASH.html")
    else:
        flash('ERROR', 'warning')

#下載位移過的檔案
@app.route("/download_HASH/<excel_name>")
def downloadfile_HASH(excel_name):
    try:
        return send_from_directory(app.config["EXCEL_HASH"], path=excel_name, as_attachment=True)
    except FileNotFoundError:
        abort(404)
#原本的程式碼return send_from_directory(app.config["CLIENT_EXCELS"], filename=excel_name, as_attachment=True)，現在filename要改成path

#下載雜湊值
@app.route("/download_SHA/<excel_name>")
def downloadfile_SHA(excel_name):
    try:
        return send_from_directory(app.config["EXCEL_SHA"], path=excel_name, as_attachment=True)
    except FileNotFoundError:
        abort(404)



app.config["EXCEL_TM"] = "/app/app/static/upload/excel_hide/excel_TM" #藏入商標後的檔案 存檔路徑

#輸入欄位名稱_藏入商標
@app.route("/add-trademark", methods=["POST"])
def add_trademark():
    return render_template("public/addTM.html")

#藏入商標
@app.route("/trademark", methods=["GET", "POST"])
def trademark():

    new_excel_tm = app.config["EXCEL_TM"] + '/processed.xlsx' #使用者上傳後的檔案

    # 使用openpyxl建立新活頁簿wb_new
    wb_new = Workbook()
    wb_new.save(new_excel_tm)

    # 使用openpyxl讀取原始檔案
    wb = load_workbook(app.config["EXCEL_UPLOADS"] + '/ori.xlsx')
    ws = wb.worksheets[0]

    # 使用openpyxl讀取new_excel
    wb_new = load_workbook(new_excel_tm)
    ws_new = wb_new.active

    a=pd.read_excel(app.config["EXCEL_UPLOADS"] + '/ori.xlsx') #讀取原檔
    df = pd.DataFrame(a)
    sp = request.args.get('TMcolname') #request.args.get('TMcolname')是由使用者輸入的欄位名稱
    List= df[sp].tolist()

    recounted = Counter(List)   #統計資料出現次數

    b1=df.groupby(sp).size()
    A=max(b1)  #找出最大值的點為多少

    a=np.array(List)   #將資料改成陣列  (分數)
    a1=np.unique(a)    

    b=np.array(b1)     #將資料轉成陣列  (次數)

    c=np.vstack((b,a1))   #合併變成二維陣列

    MAX=max(List)
    #plt.hist(List,range(0,MAX+3),align='left', edgecolor='#000000',linewidth=2)
    #plt.xticks(List)
    #plt.savefig("hist2-0.png")
    #plt.close()

    n=-1
    p=[]

    for row in c:    
        for col in row: 
            n=n+1
            if(col==A):     #判斷次數如果是跟最大一樣
                p.append(c[1][n])

    M=p[0]     #找第一個出現的最大值
    #print("最高點",M)
    m=-1
    o=[]
    for i in a:
        m=m+1  
        if i < M:  
            i=i-1
            df.at[m,sp] = i
            DataFrame(df).to_excel(new_excel_tm, sheet_name='Sheet1', index=False, header=True)
            if(i<0): 
                o.append(m)
        else:     
            if(i==M):
                df.at[m,sp] = i
                DataFrame(df).to_excel(new_excel_tm, sheet_name='Sheet1', index=False, header=True)
            else:
                i=i+1
                df.at[m,sp] = i
                DataFrame(df).to_excel(new_excel_tm, sheet_name='Sheet1', index=False, header=True)  

    def union_without_repetition(list1,list2):
        result = list(set(list1) | set(list2))
        return result

    r=pd.read_excel(new_excel_tm, sheet_name='Sheet1')
    sr= df[sp].tolist()
    ar=np.array(sr)

    bins_listx=union_without_repetition(List,sr)
    #plt.hist(sr,range(-1,MAX+3),align='left', edgecolor='#000000',linewidth=2)
    #plt.xticks(bins_listx)
    #plt.savefig("hist2-1.png")
    #plt.close()

    sy = request.args.get('trademark')
    p= ' '.join(format(ord(c), 'b') for c in sy)
    #print("藏入資料:",p)
    tc=len(p)
    #print("長度:",tc)

    n=-1
    t=-1
    for i in ar:
        n=n+1
        if(i==M):
            t=t+1    #計算第在第幾個資料
            if(t<tc):
                if(p[t]=='1'):
                    i=i-1
                    df.at[n,sp] = i
                    DataFrame(df).to_excel(new_excel_tm, sheet_name='Sheet1', index=False, header=True)
                if(p[t]=='0'):
                    DataFrame(df).to_excel(new_excel_tm, sheet_name='Sheet1', index=False, header=True)
                if(p[t]==' '):
                    i=i+1
                    df.at[n,sp] = i
                    DataFrame(df).to_excel(new_excel_tm, sheet_name='Sheet1', index=False, header=True)
            else:
                t=-1
                i=i+1
                df.at[n,sp]=i
                DataFrame(df).to_excel(new_excel_tm, sheet_name='Sheet1', index=False, header=True)
        else:
            df.at[n,sp]=i
            DataFrame(df).to_excel(new_excel_tm, sheet_name='Sheet1', index=False, header=True)


    def union_without_repetition(list1,list2):
        result = list(set(list1) | set(list2))
        return result

    wb=load_workbook(new_excel_tm)
    sheet_ranges = wb['Sheet1']

    k = pd.read_excel(new_excel_tm, nrows=0)
    L = k.columns.tolist()
    r=L.index(sp)+1

    font = Font(size=15)

    for i in o:
        i=i+2 #PYTHON讀的格子0 在EXCEL是2
        f = sheet_ranges.cell(row=i, column=r)
        f.font = font
        f.value=0
        wb.save(new_excel_tm)

    r=pd.read_excel(new_excel_tm)
    sx= df[sp].tolist()
    bins_listx2=union_without_repetition(bins_listx,sx)
    #plt.hist(sx,range(-1,MAX+3),align='left', edgecolor='#000000',linewidth=2)
    #plt.xticks(bins_listx2)
    #plt.savefig("hist2-2.png")
    if M is not None:
        popup(M) #FLASH最高值
        return render_template("public/addTM.html")
    else:
        flash('ERROR', 'warning')
    

#下載位移過的檔案
@app.route("/download_TM/<excel_name>")
def downloadfile_TM(excel_name):
    try:
        return send_from_directory(app.config["EXCEL_TM"], path=excel_name, as_attachment=True)
    except FileNotFoundError:
        abort(404)
#原本的程式碼return send_from_directory(app.config["CLIENT_EXCELS"], filename=excel_name, as_attachment=True)，現在filename要改成path

app.config["EXCEL_UPLOADS_RE"] = "/app/app/static/upload/excel_re/re_ori" #還原頁面上傳的檔案(位移過的檔案) 存檔路徑
app.config["EXCEL_RE"] = "/app/app/static/upload/excel_re/re_new" #還原後檔案 存檔路徑

#上傳要還原的檔案
@app.route("/upload-excel-re", methods=["GET", "POST"])
def upload_excel_re():

    if request.method == "POST":

        if request.files:

            excel = request.files["excel_re"]

            if excel.filename == "":
                flash('未選取檔案', 'warning')
                return redirect(request.url)

            if allowed_excel(excel.filename):
                filename = secure_filename(excel.filename)
                
                #如果檔名已經存在，則刪除舊檔，建立新檔
                if os.path.isfile(app.config["EXCEL_UPLOADS_RE"] + excel.filename):
                    os.remove(app.config["EXCEL_UPLOADS_RE"] + excel.filename)
                    excel.save(os.path.join(app.config["EXCEL_UPLOADS_RE"], excel.filename))
                else:
                    excel.save(os.path.join(app.config["EXCEL_UPLOADS_RE"], excel.filename))

                #如果ori.xlsx存在，則刪除舊檔，建立新檔
                str_upload_path = str(app.config["EXCEL_UPLOADS_RE"])
                if os.path.isfile(app.config["EXCEL_UPLOADS_RE"] + "/ori_reverse.xlsx"):
                    os.remove(app.config["EXCEL_UPLOADS_RE"] + "/ori_reverse.xlsx")
                    os.rename(str_upload_path + "/" + excel.filename,str_upload_path + "/" + "ori_reverse.xlsx")
                else:
                    os.rename(str_upload_path + "/" + excel.filename,str_upload_path + "/" + "ori_reverse.xlsx")

                flash('Excel saved', 'success')
                return redirect(request.url)
                #return redirect("/download/"+filename) #會下載剛剛上傳的檔案

            else:
                flash('請上傳附檔名為 .xlsx 的檔案', 'warning')
                return redirect(request.url)

    return render_template("public/data_reversing.html")

#輸入欄位名稱和最高點_取出商標或雜湊
@app.route("/take-out-mes", methods=["GET", "POST"])
def take_out_mes():
    return render_template("public/take_out.html")

#取出商標或雜湊
@app.route("/mes", methods=["GET", "POST"])
def mes():

    M = int(request.args.get('peak')) #最高點

    re_ori_file = app.config["EXCEL_UPLOADS_RE"] + '/ori_reverse.xlsx' #還原頁面上傳的檔案(位移過的檔案) 存檔路徑
    re_file = app.config["EXCEL_RE"] + '/recovered.xlsx' #還原後檔案 存檔路徑

    a=pd.read_excel(re_ori_file) #讀原檔
    df = pd.DataFrame(a)

    sp = str(request.args.get('MEScolname')) #要修改的欄位
    List= df[sp].tolist() 

    wb=load_workbook(re_ori_file)
    sheet_ranges = wb['Sheet1']

    a=np.array(List)   #將資料改成陣列  (分數)  

    k = pd.read_excel(re_ori_file,nrows=0)
    L = k.columns.tolist()
    print(L) 
    r=L.index(sp)+1
    print(r)


    d=-1
    for k in a:
        d=d+1
        i=d+2
        f = sheet_ranges.cell(row=i, column=r)
        t=f.font
        if(t.size==15):
        # print(d,i)
            k=-1
            df.at[d,sp] = k
            DataFrame(df).to_excel(re_file,sheet_name='Sheet1', index=False, header=True) 
            
    n=-1
    for i in a:
        n=n+1
        if(i==M-1):
            i=i+1
            df.at[n,sp] = i
            DataFrame(df).to_excel(re_file,sheet_name='Sheet1', index=False, header=True)    
        if(i==M+1):
            i=i-1
            df.at[n,sp]=i
            DataFrame(df).to_excel(re_file,sheet_name='Sheet1', index=False, header=True)

    a1=pd.read_excel(re_file)
    df1 = pd.DataFrame(a1)
    List1= df1[sp].tolist()

    a2=np.array(List1)

    m=-1
    for e in a2:
        m=m+1  
        if(e < M):  
            e=e+1
            df.at[m,sp] = e
            DataFrame(df).to_excel(re_file,sheet_name='Sheet1', index=False, header=True)
        if(e > M):  
            e=e-1
            df.at[m,sp] = e
            DataFrame(df).to_excel(re_file,sheet_name='Sheet1', index=False, header=True)
        else:    
            df.at[m,sp] = e
            DataFrame(df).to_excel(re_file,sheet_name='Sheet1', index=False, header=True)
    return redirect("/download_RE" + "/recovered.xlsx")

#輸入欄位名稱和最高點_取出雜湊
@app.route("/take-out-hash", methods=["GET", "POST"])
def take_out_HASH():
    return render_template("public/takeoutHASH.html")

#取出雜湊並比對
@app.route("/HASH_RE", methods=["GET", "POST"])
def hashRE():

    M = int(request.args.get('REHASHpeak')) #最高點

    re_ori_file = app.config["EXCEL_UPLOADS_RE"] + '/ori_reverse.xlsx' #還原頁面上傳的檔案(位移過的檔案) 存檔路徑
    re_file = app.config["EXCEL_RE"] + '/recovered.xlsx' #還原後檔案 存檔路徑

    a=pd.read_excel(re_ori_file) #讀使用者上傳的檔案
    df = pd.DataFrame(a)

    sp = str(request.args.get('REHASHcolname')) #要修改的欄位
    List= df[sp].tolist() 

    wb=load_workbook(re_ori_file)
    sheet_ranges = wb['Sheet1']

    a=np.array(List)   #將資料改成陣列  (分數)  

    k = pd.read_excel(re_ori_file,nrows=0)
    L = k.columns.tolist()
    print(L) 
    r=L.index(sp)+1
    print(r)


    d=-1
    for k in a:
        d=d+1
        i=d+2
        f = sheet_ranges.cell(row=i, column=r)
        t=f.font
        if(t.size==15):
        # print(d,i)
            k=-1
            df.at[d,sp] = k
            DataFrame(df).to_excel(re_file, sheet_name='Sheet1', index=False, header=True) 
            
    n=-1
    for i in a:
        n=n+1
        if(i==M-1):
            i=i+1
            df.at[n,sp] = i
            DataFrame(df).to_excel(re_file, sheet_name='Sheet1', index=False, header=True)    
        if(i==M+1):
            i=i-1
            df.at[n,sp]=i
            DataFrame(df).to_excel(re_file, sheet_name='Sheet1', index=False, header=True)

    a1=pd.read_excel(re_file)
    df1 = pd.DataFrame(a1)
    List1= df1[sp].tolist()

    a2=np.array(List1)

    m=-1
    for e in a2:
        m=m+1  
        if(e < M):  
            e=e+1
            df.at[m,sp] = e
            DataFrame(df).to_excel(re_file, sheet_name='Sheet1', index=False, header=True)
        if(e > M):  
            e=e-1
            df.at[m,sp] = e
            DataFrame(df).to_excel(re_file, sheet_name='Sheet1', index=False, header=True)
        else:    
            df.at[m,sp] = e
            DataFrame(df).to_excel(re_file, sheet_name='Sheet1', index=False, header=True)

    path = app.config["EXCEL_RE"] + '/new_HASH.txt'
    f = open(path, 'w',encoding='UTF-8')

    a=pd.read_excel(re_file, sheet_name='Sheet1')
    df = pd.DataFrame(a)

    sp = str(request.args.get('REHASHcolname'))
    List= df[sp].tolist()

    df1=df[sp].apply(str)
    #print(df1)
    u=[]
    for i in df1:
        u.append(i)
    #print(u)

    T="".join(u)
    #print(T)

    sha = hashlib.md5(T.encode("utf-8")).hexdigest()
    print(sha)
    print(sha,file=f)  

    bin_str = ""
    for n in sha:
        bin_str += bin(int(n,16))[2:].zfill(4)

    print(bin_str) 
    print(bin_str,file=f)

    #比對雜湊值
    window = tk.Tk()
    window.title('比對驗證碼')
    window.geometry("300x80")

    ori_hash = app.config["EXCEL_RE"] + '/ori_HASH.txt' #使用者上傳後轉成雜湊
    new_hash = app.config["EXCEL_RE"] + '/new_HASH.txt' #從RE取出來的雜湊

    if filecmp.cmp(ori_hash, new_hash):
        print("檔案相同")
        label = tk.Label(window,text = '檔案沒有被修改過',font = ('Arial', 16)) 
    else:
        print("檔案不同")
        label = tk.Label(window,text = '檔案被修改過',font = ('Arial', 16)) 

    label.pack()
    window.mainloop()
            
    return render_template("public/takeoutHASH.html")            
    #return redirect("/download_RE" + "/recovered.xlsx")

#輸入欄位名稱和最高點_取出商標
@app.route("/take-out-tm", methods=["GET", "POST"])
def take_out_TM():
    return render_template("public/takeoutTM.html")

def decode(s): #由ASCII二進位刑式轉回ASCII對應字元
    return ''.join([chr(i) for i in [int(b, 2) for b in s.split()]])

#取出商標
@app.route("/TM_RE", methods=["GET", "POST"])
def tmRE():

    M = int(request.args.get('RETMpeak')) #最高點

    re_ori_file = app.config["EXCEL_UPLOADS_RE"] + '/ori_reverse.xlsx' #還原頁面上傳的檔案(位移過的檔案) 存檔路徑
    re_file = app.config["EXCEL_RE"] + '/recovered.xlsx' #還原後檔案 存檔路徑

    path = app.config["EXCEL_RE"] + '/extracted_logo.txt'
    f = open(path, 'w',encoding='UTF-8')

    a=pd.read_excel(re_ori_file) #讀原檔
    df = pd.DataFrame(a)

    sp = str(request.args.get('RETMcolname')) #要讀取的欄位
    List= df[sp].tolist()  
    n=-1
    k=""
    for i in List:
        n=n+1
        if(i==M):
            k=k+"0"
        if(i==M-1):
            k=k+"1" 
        if(i==M+1):
            k=k+" "
    #讀商標
    print(decode(k))
    print(decode(k),file=f)  
    f.close() 
            
    return redirect("download_RE" + "/extracted_logo.txt")

#下載位移過的檔案
@app.route("/download_RE/<excel_name>")
def downloadfile_RE(excel_name):
    try:
        return send_from_directory(app.config["EXCEL_RE"], path=excel_name, as_attachment=True)
    except FileNotFoundError:
        abort(404)
#原本的程式碼return send_from_directory(app.config["CLIENT_EXCELS"], filename=excel_name, as_attachment=True)，現在filename要改成path
#還原 存還原
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import xlrd
import hashlib
from pandas import DataFrame
from collections import Counter
from openpyxl import load_workbook
from openpyxl.styles  import Font
M=28
a=pd.read_excel('C:/Users/建良/Desktop/co/hide.xlsx')
df = pd.DataFrame(a)

sp = str(input("請輸入要修改欄位："))
List= df[sp].tolist() 

wb=load_workbook(filename = 'hide.xlsx')
sheet_ranges = wb['Sheet1']

a=np.array(List)   #將資料改成陣列  (分數)  

k = pd.read_excel("hide.xlsx",nrows=0)
L = k.columns.tolist()
print(L) 
r=L.index(sp)+1
print(r)


d=-1
for k in a:
    d=d+1
    i=d+2
    f = sheet_ranges.cell(row=i, column=r)
    t=f.font
    if(t.size==15):
       # print(d,i)
        k=-1
        df.at[d,sp] = k
        DataFrame(df).to_excel('C:/Users/建良/Desktop/co/hide0.xlsx',sheet_name='Sheet1', index=False, header=True) 
        
n=-1
for i in a:
    n=n+1
    if(i==M-1):
        i=i+1
        df.at[n,sp] = i
        DataFrame(df).to_excel('C:/Users/建良/Desktop/co/hide0.xlsx',sheet_name='Sheet1', index=False, header=True)    
    if(i==M+1):
        i=i-1
        df.at[n,sp]=i
        DataFrame(df).to_excel('C:/Users/建良/Desktop/co/hide0.xlsx',sheet_name='Sheet1', index=False, header=True)

a1=pd.read_excel('C:/Users/建良/Desktop/co/hide0.xlsx')
df1 = pd.DataFrame(a1)
List1= df1[sp].tolist()

a2=np.array(List1)

m=-1
for e in a2:
    m=m+1  
    if(e < M):  
        e=e+1
        df.at[m,sp] = e
        DataFrame(df).to_excel('C:/Users/建良/Desktop/co/hide0.xlsx',sheet_name='Sheet1', index=False, header=True)
    if(e > M):  
        e=e-1
        df.at[m,sp] = e
        DataFrame(df).to_excel('C:/Users/建良/Desktop/co/hide0.xlsx',sheet_name='Sheet1', index=False, header=True)
    else:    
        df.at[m,sp] = e
        DataFrame(df).to_excel('C:/Users/建良/Desktop/co/hide0.xlsx',sheet_name='Sheet1', index=False, header=True)